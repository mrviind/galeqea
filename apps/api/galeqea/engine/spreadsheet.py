"""Requirement extraction from spreadsheets.

Requirement registers arrive as spreadsheets more often than as prose - a
traceability matrix, a numbered backlog export, an RTM handed over by a client.
Treating one as flat text loses the structure that makes it valuable, so this
reads it as a table and works out which column means what.

The header is *detected*, not assumed to be row 1: real requirement workbooks
routinely open with a title, a revision block, and a blank row before the table
starts.
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field

#: Column meanings, in priority order. The first pattern that matches a header
#: cell claims that column, so more specific names must come first.
COLUMN_PATTERNS: list[tuple[str, re.Pattern[str]]] = [
    ("ref", re.compile(r"^(req(uirement)?[\s_]*(id|no|ref|#|key)|id|ref(erence)?|key|item[\s_]*(no|#))$", re.I)),
    ("title", re.compile(r"^(requirement|description|summary|title|user[\s_]*story|statement|need)$", re.I)),
    ("acceptance", re.compile(r"^(acceptance[\s_]*criteria|ac|expected[\s_]*(result|behaviou?r)|success[\s_]*criteria)$", re.I)),
    ("priority", re.compile(r"^(priority|severity|importance|moscow)$", re.I)),
    ("risk", re.compile(r"^(risk|risk[\s_]*level|criticality)$", re.I)),
    ("kind", re.compile(r"^(type|category|kind|classification|requirement[\s_]*type)$", re.I)),
    ("section", re.compile(r"^(module|component|area|feature|epic|section|category)$", re.I)),
    ("notes", re.compile(r"^(notes?|comments?|remarks?|rationale)$", re.I)),
]

#: A row whose ref cell matches this is a section heading, not a requirement.
SECTION_ROW = re.compile(r"^(section|module|epic|feature|group)\b", re.I)

MAX_SCAN_ROWS = 25
MAX_ROWS = 5000


@dataclass(slots=True)
class SheetRequirement:
    ref: str
    title: str
    text: str = ""
    section: str = ""
    kind: str = "functional"
    priority: str = ""
    risk: str = ""
    acceptance_criteria: list[str] = field(default_factory=list)
    source_row: int = 0


@dataclass(slots=True)
class SheetResult:
    requirements: list[SheetRequirement] = field(default_factory=list)
    sheets_read: list[str] = field(default_factory=list)
    columns: dict[str, str] = field(default_factory=dict)
    warnings: list[str] = field(default_factory=list)
    rows_skipped: int = 0

    def as_text(self) -> str:
        """A readable rendering, kept as the document body for provenance."""
        lines: list[str] = []
        current_section = ""
        for item in self.requirements:
            if item.section and item.section != current_section:
                current_section = item.section
                lines.append(f"\n## {current_section}")
            lines.append(f"\n{item.ref} {item.title}")
            if item.text and item.text != item.title:
                lines.append(item.text)
            for criterion in item.acceptance_criteria:
                lines.append(f"- {criterion}")
        return "\n".join(lines).strip()


def looks_like_spreadsheet(filename: str, mime_type: str = "") -> bool:
    return (
        filename.lower().endswith((".xlsx", ".xlsm", ".xls"))
        or "spreadsheetml" in mime_type
        or mime_type == "application/vnd.ms-excel"
    )


def extract(data: bytes, filename: str = "") -> SheetResult:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return SheetResult(warnings=["spreadsheet support needs `openpyxl` (pip install openpyxl)"])

    if filename.lower().endswith(".xls"):
        return SheetResult(warnings=[
            "This is a legacy .xls file, which openpyxl cannot read. Re-save it as "
            ".xlsx and upload again — no data is lost in that conversion."
        ])

    try:
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - openpyxl raises a wide variety
        return SheetResult(warnings=[f"could not open the workbook: {exc}"])

    result = SheetResult()
    counter = 0

    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(max_row=MAX_ROWS, values_only=True))
        if not rows:
            continue

        header_index, mapping = _find_header(rows)
        if header_index is None:
            result.warnings.append(
                f"sheet '{sheet.title}' has no recognisable header row "
                "(looked for a column named ID/Ref and one named Requirement/Description)"
            )
            continue

        result.sheets_read.append(sheet.title)
        result.columns.update({v: k for k, v in mapping.items()})
        current_section = ""

        for row_number, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
            cells = {name: _cell(row, index) for name, index in mapping.items()}
            ref = cells.get("ref", "")
            title = cells.get("title", "")

            if not ref and not title:
                continue

            # A row with a label but no requirement text is a section banner.
            if ref and not title and SECTION_ROW.match(ref):
                current_section = ref
                continue
            if not title:
                result.rows_skipped += 1
                continue

            if not ref:
                counter += 1
                ref = f"REQ-{counter:03d}"

            criteria = _split_criteria(cells.get("acceptance", ""))
            result.requirements.append(SheetRequirement(
                ref=ref.strip()[:64],
                title=title.strip()[:400],
                text=" ".join(filter(None, [title, cells.get("notes", "")])).strip()[:4000],
                section=(cells.get("section") or current_section)[:200],
                kind=_normalise_kind(cells.get("kind", "")),
                priority=_normalise_priority(cells.get("priority", "")),
                risk=_normalise_risk(cells.get("risk", "") or cells.get("priority", "")),
                acceptance_criteria=criteria,
                source_row=row_number,
            ))

    workbook.close()

    if not result.requirements and not result.warnings:
        result.warnings.append("no requirement rows were found in this workbook")
    return result


# --------------------------------------------------------------------------- #
def _find_header(rows: list[tuple]) -> tuple[int | None, dict[str, int]]:
    """Locate the header row and map its columns to meanings.

    Scored rather than taken on faith: a workbook that opens with a title block
    would otherwise have its first line read as headers, and every requirement
    would come out mis-columned.
    """
    best: tuple[int, dict[str, int], int] | None = None

    for index, row in enumerate(rows[:MAX_SCAN_ROWS]):
        mapping: dict[str, int] = {}
        for column, value in enumerate(row):
            if not isinstance(value, str):
                continue
            header = value.strip()
            if not header or len(header) > 60:
                continue
            for name, pattern in COLUMN_PATTERNS:
                if name not in mapping and pattern.match(header):
                    mapping[name] = column
                    break
        # A header row must offer something to name a requirement by.
        if "title" not in mapping and "ref" not in mapping:
            continue
        score = len(mapping) + (2 if "title" in mapping else 0)
        if best is None or score > best[2]:
            best = (index, mapping, score)

    if best is None:
        return None, {}
    return best[0], best[1]


def _cell(row: tuple, index: int) -> str:
    if index >= len(row):
        return ""
    value = row[index]
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _split_criteria(value: str) -> list[str]:
    """Acceptance criteria arrive as bullets, numbers or newlines in one cell."""
    if not value:
        return []
    parts = re.split(r"\n+|(?:^|\s)[-•*]\s+|(?:^|\s)\d+[.)]\s+", value)
    # Splitting on a newline consumes the whitespace a following bullet needed
    # to be recognised, so the marker survives on every line after the first.
    # Stripping is more robust than trying to make one regex handle both.
    cleaned = [re.sub(r"^\s*(?:[-•*]|\d+[.)])\s*", "", p).strip() for p in parts]
    return [p for p in cleaned if len(p) > 4][:10]


def _normalise_priority(value: str) -> str:
    lowered = value.strip().lower()
    if lowered in {"1", "p1", "critical", "highest", "blocker", "must", "must have"}:
        return "critical"
    if lowered in {"2", "p2", "high", "should", "should have"}:
        return "high"
    if lowered in {"3", "p3", "medium", "normal", "could", "could have"}:
        return "medium"
    if lowered in {"4", "p4", "low", "lowest", "minor", "wont", "won't have"}:
        return "low"
    return ""


def _normalise_risk(value: str) -> str:
    return _normalise_priority(value)


def _normalise_kind(value: str) -> str:
    lowered = value.strip().lower()
    if any(word in lowered for word in ("non-functional", "nfr", "performance", "security", "usability")):
        return "non_functional"
    if "story" in lowered:
        return "user_story"
    if any(word in lowered for word in ("scenario", "use case", "uc")):
        return "scenario"
    return "functional"
